from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from datetime import date, timedelta


from .models import *
from .forms import *
from employee.models import *
from accounts.decorators import hr_required, check_module_permission

@login_required
@check_module_permission('contract', 'View')
def contract_list(request):
    """List all contracts with filtering options"""
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')
    
    contracts = EmploymentContract.objects.all().order_by('-contract_id')
    
    if query:
        contracts = contracts.filter(
            Q(employee__full_name__icontains=query) |
            Q(contract_id__icontains=query)
        )
    
    if status_filter:
        contracts = contracts.filter(status=status_filter)
        
    if type_filter:
        contracts = contracts.filter(contract_type=type_filter)
    
    # Add upcoming expiration filter
    expiring_soon = request.GET.get('expiring', '')
    if expiring_soon:
        thirty_days_later = date.today() + timedelta(days=30)
        contracts = contracts.filter(
            end_date__isnull=False,
            end_date__lte=thirty_days_later,
            end_date__gte=date.today(),
            status='Active'
        )
    
    paginator = Paginator(contracts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'contract/contract_list.html', {
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'expiring_soon': expiring_soon
    })

@login_required
@check_module_permission('contract', 'Edit')
def contract_create(request):
    """Create new employment contract"""
    if request.method == 'POST':
        form = EmploymentContractForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save()
            messages.success(request, 'Contract created successfully')
            return redirect('contract_detail', pk=contract.contract_id)
    else:
        # Pre-populate employee if provided in URL
        employee_id = request.GET.get('employee_id')
        initial_data = {}
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                initial_data['employee'] = employee
            except Employee.DoesNotExist:
                pass
        
        form = EmploymentContractForm(initial=initial_data)
    
    return render(request, 'contract/contract_form.html', {'form': form})

@login_required
@check_module_permission('contract', 'View')
def contract_detail(request, pk):
    """View contract details"""
    contract = get_object_or_404(EmploymentContract, pk=pk)
    return render(request, 'contract/contract_detail.html', {'contract': contract})

@login_required
@hr_required
def contract_terminate(request, pk):
    """Terminate an active contract"""
    contract = get_object_or_404(EmploymentContract, pk=pk)
    
    if contract.status != 'Active':
        messages.error(request, 'Only active contracts can be terminated')
        return redirect('contract_detail', pk=pk)
    
    if request.method == 'POST':
        termination_reason = request.POST.get('termination_reason')
        contract.status = 'Terminated'
        contract.notes = f"{contract.notes}\n\nTerminated on {date.today()}: {termination_reason}"
        contract.save()
        
        # Update employee status if necessary
        employee = contract.employee
        active_contracts = EmploymentContract.objects.filter(
            employee=employee,
            status='Active'
        ).exists()
        
        if not active_contracts:
            employee.status = 'Resigned'
            employee.save()
        
        messages.success(request, 'Contract terminated successfully')
        return redirect('contract_detail', pk=pk)
    
    return render(request, 'contract/contract_terminate.html', {'contract': contract})


@login_required
@check_module_permission('contract', 'Edit')
def contract_update(request, pk):
    """Update an existing employment contract"""
    contract = get_object_or_404(EmploymentContract, pk=pk)
    
    # Check if contract can be edited
    if contract.status != 'Active':
        messages.warning(request, f'This contract is {contract.status} and some changes may be restricted')
    
    if request.method == 'POST':
        form = EmploymentContractForm(request.POST, request.FILES, instance=contract)
        if form.is_valid():
            updated_contract = form.save()
            messages.success(request, 'Contract updated successfully')
            return redirect('contract_detail', pk=updated_contract.contract_id)
    else:
        form = EmploymentContractForm(instance=contract)
    
    return render(request, 'contract/contract_form.html', {
        'form': form,
        'contract': contract,
        'is_update': True
    })

@login_required
@hr_required
def contract_renew(request, pk):
    """Renew an existing employment contract"""
    contract = get_object_or_404(EmploymentContract, pk=pk)
    
    # Only allow renewal for contracts that are about to expire or have expired
    if contract.status not in ['Active', 'Expired']:
        messages.error(request, f'Contracts with status "{contract.status}" cannot be renewed')
        return redirect('contract_detail', pk=pk)
    
    if request.method == 'POST':
        form = ContractRenewalForm(request.POST)
        if form.is_valid():
            # Mark the current contract as expired
            contract.status = 'Expired'
            contract.notes = f"{contract.notes or ''}\n\nRenewed on {date.today()}"
            contract.save()
            
            # Create a new contract
            new_contract = EmploymentContract.objects.create(
                employee=contract.employee,
                contract_type=form.cleaned_data['new_contract_type'],
                start_date=form.cleaned_data['new_start_date'],
                end_date=form.cleaned_data['new_end_date'],
                base_salary=form.cleaned_data['new_base_salary'],
                allowance=form.cleaned_data['new_allowance'],
                signed_by=request.user.get_full_name() or request.user.username,
                sign_date=date.today(),
                notes=f"Renewal of contract #{contract.contract_id}. {form.cleaned_data['renewal_notes'] or ''}",
                status='Active'
            )
            
            messages.success(request, 'Contract renewed successfully')
            return redirect('contract_detail', pk=new_contract.contract_id)
    else:
        # Initialize form with values from the current contract
        initial_data = {
            'new_start_date': date.today(),
            'new_end_date': None if contract.contract_type == 'Indefinite-term' else (contract.end_date + timedelta(days=365) if contract.end_date else None),
            'new_contract_type': contract.contract_type,
            'new_base_salary': contract.base_salary,
            'new_allowance': contract.allowance,
        }
        form = ContractRenewalForm(initial=initial_data)
    
    return render(request, 'contract/contract_renew.html', {
        'form': form,
        'contract': contract
    })

@login_required
@check_module_permission('contract', 'View')
def employee_contracts(request, employee_id):
    """View all contracts for a specific employee"""
    employee = get_object_or_404(Employee, pk=employee_id)
    
    contracts = EmploymentContract.objects.filter(
        employee=employee
    ).order_by('-start_date')
    
    return render(request, 'contract/employee_contracts.html', {
        'employee': employee,
        'contracts': contracts,
    })

@login_required
@check_module_permission('contract', 'Edit')
def employee_contract_create(request, employee_id):
    """Create a new contract for a specific employee"""
    employee = get_object_or_404(Employee, pk=employee_id)
    
    if request.method == 'POST':
        form = EmploymentContractForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save()
            messages.success(request, 'Contract created successfully')
            return redirect('employee_contracts', employee_id=employee_id)
    else:
        form = EmploymentContractForm(initial={
            'employee': employee,
            'start_date': date.today(),
            'sign_date': date.today()
        })
        # Lock the employee field since we're creating for a specific employee
        form.fields['employee'].widget.attrs['readonly'] = True
    
    return render(request, 'contract/contract_form.html', {
        'form': form,
        'employee': employee
    })

@login_required
@hr_required
def export_contracts(request):
    """Export contracts data as CSV or Excel"""
    from django.http import HttpResponse
    import csv
    from openpyxl import Workbook
    
    format_type = request.GET.get('format', 'csv')
    status_filter = request.GET.get('status', '')
    
    # Get contracts with filtering
    contracts = EmploymentContract.objects.all().select_related('employee')
    
    if status_filter:
        contracts = contracts.filter(status=status_filter)
    
    # Fields to export
    fields = [
        'contract_id', 'employee__full_name', 'contract_type', 'status',
        'start_date', 'end_date', 'base_salary', 'allowance', 'sign_date'
    ]
    
    headers = [
        'Contract ID', 'Employee Name', 'Contract Type', 'Status',
        'Start Date', 'End Date', 'Base Salary', 'Allowance', 'Signed Date'
    ]
    
    if format_type == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Contracts"
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data rows
        for row_num, contract in enumerate(contracts, 2):
            ws.cell(row=row_num, column=1, value=contract.contract_id)
            ws.cell(row=row_num, column=2, value=contract.employee.full_name)
            ws.cell(row=row_num, column=3, value=contract.contract_type)
            ws.cell(row=row_num, column=4, value=contract.status)
            ws.cell(row=row_num, column=5, value=contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '')
            ws.cell(row=row_num, column=6, value=contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '')
            ws.cell(row=row_num, column=7, value=float(contract.base_salary))
            ws.cell(row=row_num, column=8, value=float(contract.allowance))
            ws.cell(row=row_num, column=9, value=contract.sign_date.strftime('%Y-%m-%d') if contract.sign_date else '')
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="contracts_export.xlsx"'
        wb.save(response)
        return response
    else:
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="contracts_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for contract in contracts:
            writer.writerow([
                contract.contract_id,
                contract.employee.full_name,
                contract.contract_type,
                contract.status,
                contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '',
                contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '',
                contract.base_salary,
                contract.allowance,
                contract.sign_date.strftime('%Y-%m-%d') if contract.sign_date else ''
            ])
        
        return response

@login_required
@check_module_permission('contract', 'View')
def expiring_contracts(request):
    """View contracts that are expiring soon"""
    days = int(request.GET.get('days', 30))
    today = date.today()
    expiry_date = today + timedelta(days=days)
    
    expiring = EmploymentContract.objects.filter(
        status='Active',
        end_date__isnull=False,
        end_date__gte=today,
        end_date__lte=expiry_date
    ).order_by('end_date')
    
    return render(request, 'contract/expiring_contracts.html', {
        'expiring_contracts': expiring,
        'days': days,
        'today': today,
        'expiry_date': expiry_date
    })

@login_required
@check_module_permission('contract', 'View')
def contract_report(request):
    """Generate a contract report with statistics"""
    today = date.today()
    contracts = EmploymentContract.objects.all()
    
    # Status counts
    status_counts = {
        'active': contracts.filter(status='Active').count(),
        'expired': contracts.filter(status='Expired').count(),
        'terminated': contracts.filter(status='Terminated').count(),
        'total': contracts.count()
    }
    
    # Contract type counts
    type_counts = {}
    for contract_type, _ in EmploymentContract.CONTRACT_TYPE_CHOICES:
        type_counts[contract_type] = contracts.filter(
            contract_type=contract_type, 
            status='Active'
        ).count()
    
    # Expiring contracts counts
    expiring_30 = contracts.filter(
        status='Active',
        end_date__isnull=False,
        end_date__gte=today,
        end_date__lte=today + timedelta(days=30)
    ).count()
    
    expiring_60 = contracts.filter(
        status='Active',
        end_date__isnull=False,
        end_date__gte=today + timedelta(days=31),
        end_date__lte=today + timedelta(days=60)
    ).count()
    
    expiring_90 = contracts.filter(
        status='Active',
        end_date__isnull=False,
        end_date__gte=today + timedelta(days=61),
        end_date__lte=today + timedelta(days=90)
    ).count()
    
    # Recently signed contracts
    recent_contracts = contracts.filter(
        sign_date__gte=today - timedelta(days=30)
    ).order_by('-sign_date')[:10]
    
    return render(request, 'contract/contract_report.html', {
        'status_counts': status_counts,
        'type_counts': type_counts,
        'expiring_30': expiring_30,
        'expiring_60': expiring_60,
        'expiring_90': expiring_90,
        'recent_contracts': recent_contracts
    })

@login_required
def my_contracts(request):
    """View an employee's own contracts"""
    if not request.user.employee:
        messages.error(request, "You don't have an employee profile.")
        return redirect('dashboard')
    
    contracts = EmploymentContract.objects.filter(
        employee=request.user.employee
    ).order_by('-start_date')
    
    return render(request, 'contract/my_contracts.html', {
        'contracts': contracts
    })